#!/usr/bin/env python3
"""
Check soil data quality for CSV, JSON, or Excel files.

Usage:
  python soil_quality_check.py <file-path>

Checks for:
- Duplicate sample rows
- High null percentage columns
- Outliers using agronomic ranges
- Invalid pH values
- Nutrient imbalances
- Suspect readings (negative values, impossible combinations)
- Texture inconsistencies (sand + silt + clay != 100%)
"""

import json
import sys
from pathlib import Path

# Agronomic ranges for outlier detection (typical agricultural soils)
AGRONOMIC_RANGES = {
    'ph': {'min': 3.5, 'max': 10.0, 'typical_min': 4.5, 'typical_max': 8.5},
    'nitrogen': {'min': 0, 'max': 500, 'typical_min': 5, 'typical_max': 100},
    'phosphorus': {'min': 0, 'max': 500, 'typical_min': 3, 'typical_max': 150},
    'potassium': {'min': 0, 'max': 1500, 'typical_min': 50, 'typical_max': 500},
    'organic_matter': {'min': 0, 'max': 100, 'typical_min': 0.5, 'typical_max': 15},
    'cec': {'min': 0, 'max': 100, 'typical_min': 2, 'typical_max': 50},
    'ec': {'min': 0, 'max': 50, 'typical_min': 0, 'typical_max': 8},
    'calcium': {'min': 0, 'max': 20000, 'typical_min': 200, 'typical_max': 10000},
    'magnesium': {'min': 0, 'max': 5000, 'typical_min': 25, 'typical_max': 1500},
    'sand': {'min': 0, 'max': 100},
    'silt': {'min': 0, 'max': 100},
    'clay': {'min': 0, 'max': 100},
}

# Ideal nutrient ratios
NUTRIENT_RATIOS = {
    'ca_mg': {'min': 2, 'max': 10, 'ideal': 5},  # Ca:Mg ratio
    'k_mg': {'min': 0.1, 'max': 0.5, 'ideal': 0.25},  # K:Mg ratio
}

SOIL_PARAMS = {
    'ph': ['ph', 'soil_ph', 'ph_water', 'ph_cacl2', 'ph_h2o'],
    'nitrogen': ['n', 'nitrogen', 'no3', 'nh4', 'total_n', 'nitrate'],
    'phosphorus': ['p', 'phosphorus', 'olsen_p', 'bray_p', 'mehlich_p'],
    'potassium': ['k', 'potassium', 'exchangeable_k'],
    'organic_matter': ['om', 'organic_matter', 'som', 'organic_c'],
    'cec': ['cec', 'cation_exchange'],
    'ec': ['ec', 'electrical_conductivity', 'salinity'],
    'calcium': ['ca', 'calcium'],
    'magnesium': ['mg', 'magnesium'],
    'sand': ['sand', 'sand_pct'],
    'silt': ['silt', 'silt_pct'],
    'clay': ['clay', 'clay_pct'],
}


def detect_soil_param(col_name):
    """Detect if column is a known soil parameter."""
    col_lower = col_name.lower().strip().replace(' ', '_').replace('-', '_')
    for param, names in SOIL_PARAMS.items():
        if col_lower in names or any(name in col_lower for name in names):
            return param
    return None


def parse_csv(file_path):
    """Parse CSV file into list of dicts."""
    rows = []
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    if not lines:
        return [], []

    header = [h.strip().strip('"') for h in lines[0].strip().split(',')]

    for line in lines[1:]:
        if line.strip():
            values = [v.strip().strip('"') for v in line.strip().split(',')]
            row = dict(zip(header, values))
            rows.append(row)

    return header, rows


def parse_json(file_path):
    """Parse JSON file into list of dicts."""
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, list) and data:
        header = list(data[0].keys()) if isinstance(data[0], dict) else []
        return header, data

    return [], []


def parse_excel(file_path):
    """Parse Excel file into list of dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, "Error: openpyxl not installed. Run: pip install openpyxl"

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)

    if not header_row:
        return [], []

    header = [str(h) if h else f"col_{i}" for i, h in enumerate(header_row)]
    rows = [dict(zip(header, row)) for row in rows_iter]
    wb.close()
    return header, rows


def find_duplicates(rows, header):
    """Find duplicate rows."""
    seen = set()
    duplicates = []

    for i, row in enumerate(rows):
        key = tuple(str(row.get(col, "")) for col in header)
        if key in seen:
            duplicates.append(i + 1)
        seen.add(key)

    return duplicates


def find_high_null_columns(rows, header, threshold=0.5):
    """Find columns with high null percentage."""
    issues = []

    for col in header:
        values = [row.get(col) for row in rows]
        null_count = sum(1 for v in values if v is None or v == "")
        null_pct = null_count / len(values) if values else 0

        if null_pct >= threshold:
            issues.append({
                "column": col,
                "null_count": null_count,
                "null_pct": round(null_pct * 100, 1)
            })

    return issues


def find_agronomic_outliers(rows, header):
    """Find outliers using agronomic ranges."""
    issues = []

    for col in header:
        param = detect_soil_param(col)
        if not param or param not in AGRONOMIC_RANGES:
            continue

        ranges = AGRONOMIC_RANGES[param]
        outliers = []
        atypical = []

        for i, row in enumerate(rows):
            try:
                value = float(str(row.get(col, "")).replace(",", ""))
                
                # Check impossible values
                if value < ranges['min'] or value > ranges['max']:
                    outliers.append({'row': i + 1, 'value': value})
                # Check atypical values
                elif 'typical_min' in ranges and 'typical_max' in ranges:
                    if value < ranges['typical_min'] or value > ranges['typical_max']:
                        atypical.append({'row': i + 1, 'value': value})
                        
            except (ValueError, TypeError):
                pass

        if outliers:
            issues.append({
                'column': col,
                'param': param,
                'type': 'impossible',
                'count': len(outliers),
                'examples': outliers[:5],
                'valid_range': f"{ranges['min']} - {ranges['max']}"
            })
        
        if atypical and len(atypical) > len(rows) * 0.1:  # More than 10% atypical
            issues.append({
                'column': col,
                'param': param,
                'type': 'atypical',
                'count': len(atypical),
                'typical_range': f"{ranges.get('typical_min', 'N/A')} - {ranges.get('typical_max', 'N/A')}"
            })

    return issues


def check_negative_values(rows, header):
    """Check for negative values in soil parameters that should always be positive."""
    issues = []

    for col in header:
        param = detect_soil_param(col)
        if not param:
            continue

        negatives = []
        for i, row in enumerate(rows):
            try:
                value = float(str(row.get(col, "")).replace(",", ""))
                if value < 0:
                    negatives.append({'row': i + 1, 'value': value})
            except (ValueError, TypeError):
                pass

        if negatives:
            issues.append({
                'column': col,
                'param': param,
                'count': len(negatives),
                'examples': negatives[:5]
            })

    return issues


def check_texture_consistency(rows, header):
    """Check if sand + silt + clay = 100% (within tolerance)."""
    sand_col = silt_col = clay_col = None
    
    for col in header:
        param = detect_soil_param(col)
        if param == 'sand':
            sand_col = col
        elif param == 'silt':
            silt_col = col
        elif param == 'clay':
            clay_col = col

    if not (sand_col and silt_col and clay_col):
        return []

    issues = []
    tolerance = 5  # Allow 5% deviation

    for i, row in enumerate(rows):
        try:
            sand = float(str(row.get(sand_col, "")).replace(",", ""))
            silt = float(str(row.get(silt_col, "")).replace(",", ""))
            clay = float(str(row.get(clay_col, "")).replace(",", ""))
            total = sand + silt + clay

            if abs(total - 100) > tolerance:
                issues.append({
                    'row': i + 1,
                    'sand': sand,
                    'silt': silt,
                    'clay': clay,
                    'total': round(total, 1)
                })
        except (ValueError, TypeError):
            pass

    return issues


def check_nutrient_ratios(rows, header):
    """Check for nutrient imbalances based on common ratios."""
    ca_col = mg_col = k_col = None
    
    for col in header:
        param = detect_soil_param(col)
        if param == 'calcium':
            ca_col = col
        elif param == 'magnesium':
            mg_col = col
        elif param == 'potassium':
            k_col = col

    issues = []

    # Check Ca:Mg ratio
    if ca_col and mg_col:
        imbalanced = []
        for i, row in enumerate(rows):
            try:
                ca = float(str(row.get(ca_col, "")).replace(",", ""))
                mg = float(str(row.get(mg_col, "")).replace(",", ""))
                if mg > 0:
                    ratio = ca / mg
                    ranges = NUTRIENT_RATIOS['ca_mg']
                    if ratio < ranges['min'] or ratio > ranges['max']:
                        imbalanced.append({'row': i + 1, 'ratio': round(ratio, 2)})
            except (ValueError, TypeError):
                pass

        if imbalanced:
            issues.append({
                'ratio': 'Ca:Mg',
                'count': len(imbalanced),
                'ideal_range': f"{NUTRIENT_RATIOS['ca_mg']['min']} - {NUTRIENT_RATIOS['ca_mg']['max']}",
                'examples': imbalanced[:5]
            })

    return issues


def quality_check(file_path):
    """Run quality checks on soil data file."""
    path = Path(file_path)

    if not path.exists():
        return f"Error: File not found: {file_path}"

    ext = path.suffix.lower()
    if ext == '.csv':
        header, rows = parse_csv(file_path)
    elif ext == '.json':
        header, rows = parse_json(file_path)
    elif ext in ('.xlsx', '.xlsm'):
        result = parse_excel(file_path)
        if isinstance(result[1], str):
            return result[1]
        header, rows = result
    else:
        return f"Error: Unsupported file type: {ext}"

    if not rows:
        return "Error: No data found in file"

    output = []
    output.append(f"## Soil Data Quality Report: {path.name}")
    output.append(f"- **Total samples:** {len(rows)}")
    output.append("")

    issues_found = False

    # Check duplicates
    duplicates = find_duplicates(rows, header)
    if duplicates:
        issues_found = True
        output.append("### ⚠️ Duplicate Samples")
        output.append(f"Found {len(duplicates)} duplicate row(s)")
        if len(duplicates) <= 10:
            output.append(f"Row numbers: {', '.join(map(str, duplicates))}")
        output.append("")

    # Check high null columns
    high_null = find_high_null_columns(rows, header)
    if high_null:
        issues_found = True
        output.append("### ⚠️ High Missing Data (>50%)")
        for issue in high_null:
            output.append(f"- **{issue['column']}**: {issue['null_count']} missing ({issue['null_pct']}%)")
        output.append("")

    # Check negative values
    negatives = check_negative_values(rows, header)
    if negatives:
        issues_found = True
        output.append("### ❌ Negative Values (Invalid)")
        for issue in negatives:
            output.append(f"- **{issue['column']}**: {issue['count']} negative value(s)")
            examples = [f"row {e['row']}: {e['value']}" for e in issue['examples']]
            output.append(f"  Examples: {', '.join(examples)}")
        output.append("")

    # Check agronomic outliers
    outliers = find_agronomic_outliers(rows, header)
    if outliers:
        issues_found = True
        output.append("### ⚠️ Outliers & Atypical Values")
        for issue in outliers:
            if issue['type'] == 'impossible':
                output.append(f"- **{issue['column']}** ({issue['param']}): {issue['count']} impossible value(s)")
                output.append(f"  Valid range: {issue['valid_range']}")
            else:
                output.append(f"- **{issue['column']}** ({issue['param']}): {issue['count']} atypical value(s)")
                output.append(f"  Typical range: {issue['typical_range']}")
        output.append("")

    # Check texture consistency
    texture_issues = check_texture_consistency(rows, header)
    if texture_issues:
        issues_found = True
        output.append("### ⚠️ Texture Inconsistencies")
        output.append(f"Found {len(texture_issues)} sample(s) where sand + silt + clay ≠ 100%")
        for issue in texture_issues[:5]:
            output.append(f"- Row {issue['row']}: {issue['sand']}% + {issue['silt']}% + {issue['clay']}% = {issue['total']}%")
        output.append("")

    # Check nutrient ratios
    ratio_issues = check_nutrient_ratios(rows, header)
    if ratio_issues:
        issues_found = True
        output.append("### ⚠️ Nutrient Imbalances")
        for issue in ratio_issues:
            output.append(f"- **{issue['ratio']} ratio**: {issue['count']} sample(s) outside ideal range ({issue['ideal_range']})")
        output.append("")

    if not issues_found:
        output.append("### ✅ No Issues Found")
        output.append("All soil data quality checks passed.")

    return "\n".join(output)


def main():
    if len(sys.argv) < 2:
        print("Usage: python soil_quality_check.py <file-path>", file=sys.stderr)
        sys.exit(1)

    result = quality_check(sys.argv[1])
    print(result)


if __name__ == "__main__":
    main()
