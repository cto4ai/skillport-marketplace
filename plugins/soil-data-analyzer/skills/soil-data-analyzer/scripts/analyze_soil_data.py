#!/usr/bin/env python3
"""
Analyze soil test data files (CSV, JSON, or Excel).

Usage:
  python analyze_soil_data.py <file-path>

Outputs statistical summary including:
- Row/column counts
- Column types with soil parameter detection
- Missing values
- Basic statistics for numeric columns
- Soil-specific assessments (pH, nutrients, etc.)
- Sample rows

v1.0.0: Initial release with soil-specific analytics
"""

import json
import sys
from pathlib import Path

# Soil parameter detection patterns (lowercase)
SOIL_PARAMS = {
    'ph': {'names': ['ph', 'soil_ph', 'ph_water', 'ph_cacl2', 'ph_h2o'], 'unit': '-'},
    'nitrogen': {'names': ['n', 'nitrogen', 'no3', 'nh4', 'total_n', 'nitrate', 'ammonium'], 'unit': 'ppm'},
    'phosphorus': {'names': ['p', 'phosphorus', 'olsen_p', 'bray_p', 'mehlich_p', 'available_p'], 'unit': 'ppm'},
    'potassium': {'names': ['k', 'potassium', 'exchangeable_k'], 'unit': 'ppm'},
    'organic_matter': {'names': ['om', 'organic_matter', 'som', 'organic_c', 'oc'], 'unit': '%'},
    'cec': {'names': ['cec', 'cation_exchange', 'cation_exchange_capacity'], 'unit': 'meq/100g'},
    'ec': {'names': ['ec', 'electrical_conductivity', 'salinity', 'conductivity'], 'unit': 'dS/m'},
    'calcium': {'names': ['ca', 'calcium'], 'unit': 'ppm'},
    'magnesium': {'names': ['mg', 'magnesium'], 'unit': 'ppm'},
    'sulfur': {'names': ['s', 'sulfur', 'so4', 'sulfate'], 'unit': 'ppm'},
    'iron': {'names': ['fe', 'iron'], 'unit': 'ppm'},
    'manganese': {'names': ['mn', 'manganese'], 'unit': 'ppm'},
    'zinc': {'names': ['zn', 'zinc'], 'unit': 'ppm'},
    'copper': {'names': ['cu', 'copper'], 'unit': 'ppm'},
    'boron': {'names': ['b', 'boron'], 'unit': 'ppm'},
    'sand': {'names': ['sand', 'sand_pct'], 'unit': '%'},
    'silt': {'names': ['silt', 'silt_pct'], 'unit': '%'},
    'clay': {'names': ['clay', 'clay_pct'], 'unit': '%'},
}

# Optimal ranges for general agriculture (can vary by crop)
OPTIMAL_RANGES = {
    'ph': {'low': 6.0, 'high': 7.0, 'critical_low': 5.0, 'critical_high': 8.5},
    'nitrogen': {'low': 20, 'high': 50, 'critical_low': 10},
    'phosphorus': {'low': 15, 'high': 50, 'critical_low': 5},
    'potassium': {'low': 120, 'high': 250, 'critical_low': 60},
    'organic_matter': {'low': 2.0, 'high': 5.0, 'critical_low': 1.0},
    'cec': {'low': 10, 'high': 25},
    'ec': {'high': 2.0, 'critical_high': 4.0},  # Lower is generally better
    'calcium': {'low': 500, 'high': 2500},
    'magnesium': {'low': 50, 'high': 300},
}


def detect_soil_param(col_name):
    """Detect if column is a known soil parameter."""
    col_lower = col_name.lower().strip().replace(' ', '_').replace('-', '_')
    for param, info in SOIL_PARAMS.items():
        if col_lower in info['names'] or any(name in col_lower for name in info['names']):
            return param, info['unit']
    return None, None


def classify_ph(value):
    """Classify pH value."""
    if value < 4.5:
        return 'extremely acidic'
    elif value < 5.5:
        return 'strongly acidic'
    elif value < 6.0:
        return 'moderately acidic'
    elif value < 6.5:
        return 'slightly acidic'
    elif value < 7.0:
        return 'neutral (slightly acidic)'
    elif value < 7.5:
        return 'neutral (slightly alkaline)'
    elif value < 8.0:
        return 'slightly alkaline'
    elif value < 8.5:
        return 'moderately alkaline'
    else:
        return 'strongly alkaline'


def assess_nutrient_level(param, value):
    """Assess nutrient level against optimal ranges."""
    if param not in OPTIMAL_RANGES:
        return 'N/A'
    
    ranges = OPTIMAL_RANGES[param]
    
    if 'critical_low' in ranges and value < ranges['critical_low']:
        return 'deficient'
    elif 'low' in ranges and value < ranges['low']:
        return 'low'
    elif 'critical_high' in ranges and value > ranges['critical_high']:
        return 'excessive'
    elif 'high' in ranges and value > ranges['high']:
        return 'high'
    else:
        return 'adequate'


def detect_type(values):
    """Detect column type from sample values."""
    non_null = [v for v in values if v is not None and v != ""]
    if not non_null:
        return "empty"

    numeric_count = 0
    for v in non_null[:100]:
        try:
            float(str(v).replace(",", ""))
            numeric_count += 1
        except (ValueError, TypeError):
            pass

    if numeric_count > len(non_null[:100]) * 0.8:
        return "numeric"
    return "string"


def parse_csv(file_path):
    """Parse CSV file into list of dicts."""
    rows = []
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    if not lines:
        return [], []

    header = [h.strip().strip('"') for h in lines[0].strip().split(',')]

    for line in lines[1:]:
        if line.strip():
            values = [v.strip().strip('"') for v in line.strip().split(',')]
            row = dict(zip(header, values))
            rows.append(row)

    return header, rows


def parse_json(file_path):
    """Parse JSON file into list of dicts."""
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, list) and data:
        header = list(data[0].keys()) if isinstance(data[0], dict) else []
        return header, data
    elif isinstance(data, dict):
        return list(data.keys()), [data]

    return [], []


def parse_excel(file_path):
    """Parse Excel file into list of dicts. Requires openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, "Error: openpyxl not installed. Run: pip install openpyxl"

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)

    if not header_row:
        return [], []

    header = [str(h) if h else f"col_{i}" for i, h in enumerate(header_row)]
    rows = []

    for row in rows_iter:
        row_dict = dict(zip(header, row))
        rows.append(row_dict)

    wb.close()
    return header, rows


def calculate_stats(values):
    """Calculate basic statistics for numeric values."""
    nums = []
    for v in values:
        try:
            nums.append(float(str(v).replace(",", "")))
        except (ValueError, TypeError):
            pass

    if not nums:
        return None

    nums.sort()
    n = len(nums)
    mean = sum(nums) / n
    variance = sum((x - mean) ** 2 for x in nums) / n
    std = variance ** 0.5

    return {
        "count": n,
        "min": round(nums[0], 2),
        "max": round(nums[-1], 2),
        "mean": round(mean, 2),
        "std": round(std, 2),
        "median": round(nums[n // 2], 2) if n % 2 == 1 else round((nums[n // 2 - 1] + nums[n // 2]) / 2, 2)
    }


def analyze_soil_data(file_path):
    """Analyze soil data file and return summary."""
    path = Path(file_path)

    if not path.exists():
        return f"Error: File not found: {file_path}"

    # Parse file
    ext = path.suffix.lower()
    if ext == '.csv':
        header, rows = parse_csv(file_path)
    elif ext == '.json':
        header, rows = parse_json(file_path)
    elif ext in ('.xlsx', '.xlsm'):
        result = parse_excel(file_path)
        if isinstance(result[1], str):
            return result[1]
        header, rows = result
    else:
        return f"Error: Unsupported file type: {ext} (use .csv, .json, or .xlsx)"

    if not rows:
        return "Error: No data found in file"

    # Build summary
    output = []
    output.append(f"## Soil Data Summary: {path.name}")
    output.append(f"- **Samples:** {len(rows)}")
    output.append(f"- **Parameters:** {len(header)}")
    output.append("")

    # Detect soil parameters
    soil_columns = {}
    for col in header:
        param, unit = detect_soil_param(col)
        if param:
            soil_columns[col] = {'param': param, 'unit': unit}
    
    if soil_columns:
        output.append(f"### Detected Soil Parameters: {len(soil_columns)}")
        output.append("")

    # Column analysis
    output.append("## Column Analysis")
    output.append("")
    output.append("| Column | Type | Soil Param | Missing | Stats |")
    output.append("|--------|------|------------|---------|-------|")

    for col in header:
        values = [row.get(col) for row in rows]
        col_type = detect_type(values)
        missing = sum(1 for v in values if v is None or v == "")
        missing_pct = f"{missing} ({100*missing/len(values):.1f}%)" if values else "N/A"

        soil_info = soil_columns.get(col, {})
        param_str = soil_info.get('param', '-')

        stats_str = "-"
        if col_type == "numeric":
            stats = calculate_stats(values)
            if stats:
                stats_str = f"min={stats['min']}, max={stats['max']}, mean={stats['mean']}"

        output.append(f"| {col} | {col_type} | {param_str} | {missing_pct} | {stats_str} |")

    # Soil-specific assessments
    if soil_columns:
        output.append("")
        output.append("## Soil Parameter Assessments")
        output.append("")

        for col, info in soil_columns.items():
            param = info['param']
            values = [row.get(col) for row in rows]
            stats = calculate_stats(values)
            
            if not stats:
                continue

            output.append(f"### {col} ({param.replace('_', ' ').title()})")
            output.append("")
            
            if param == 'ph':
                ph_class = classify_ph(stats['mean'])
                output.append(f"- **Average pH:** {stats['mean']} ({ph_class})")
                output.append(f"- **Range:** {stats['min']} - {stats['max']}")
                
                # Count samples in each category
                nums = [float(str(v).replace(',', '')) for v in values if v not in (None, '')]
                acidic = sum(1 for v in nums if v < 6.0)
                neutral = sum(1 for v in nums if 6.0 <= v <= 7.5)
                alkaline = sum(1 for v in nums if v > 7.5)
                output.append(f"- **Distribution:** {acidic} acidic (<6.0), {neutral} neutral (6.0-7.5), {alkaline} alkaline (>7.5)")
                
            elif param in OPTIMAL_RANGES:
                level = assess_nutrient_level(param, stats['mean'])
                output.append(f"- **Average:** {stats['mean']} {info['unit']} ({level})")
                output.append(f"- **Range:** {stats['min']} - {stats['max']} {info['unit']}")
                
                # Count samples by level
                nums = [float(str(v).replace(',', '')) for v in values if v not in (None, '')]
                levels = {'deficient': 0, 'low': 0, 'adequate': 0, 'high': 0, 'excessive': 0}
                for v in nums:
                    lvl = assess_nutrient_level(param, v)
                    if lvl in levels:
                        levels[lvl] += 1
                
                level_parts = [f"{count} {lvl}" for lvl, count in levels.items() if count > 0]
                if level_parts:
                    output.append(f"- **Sample distribution:** {', '.join(level_parts)}")
            else:
                output.append(f"- **Average:** {stats['mean']} {info['unit']}")
                output.append(f"- **Range:** {stats['min']} - {stats['max']} {info['unit']}")
            
            output.append("")

    # Sample rows
    output.append("## Sample Data (first 5 rows)")
    output.append("")

    sample = rows[:5]
    if sample:
        output.append("| " + " | ".join(header) + " |")
        output.append("|" + "|".join(["---"] * len(header)) + "|")
        for row in sample:
            vals = [str(row.get(col, ""))[:20] for col in header]
            output.append("| " + " | ".join(vals) + " |")

    return "\n".join(output)


def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_soil_data.py <file-path>", file=sys.stderr)
        sys.exit(1)

    result = analyze_soil_data(sys.argv[1])
    print(result)


if __name__ == "__main__":
    main()
