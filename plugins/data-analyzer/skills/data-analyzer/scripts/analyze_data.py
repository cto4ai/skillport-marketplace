#!/usr/bin/env python3
"""
Analyze CSV, JSON, or Excel data files.

Usage:
  python analyze_data.py <file-path>

Outputs statistical summary including:
- Row/column counts
- Column types
- Missing values
- Basic statistics for numeric columns
- Sample rows

v1.1.0: Added Excel (.xlsx) support
v1.1.1: Added standard deviation to numeric stats
"""

import json
import sys
from pathlib import Path


def detect_type(values):
    """Detect column type from sample values."""
    non_null = [v for v in values if v is not None and v != ""]
    if not non_null:
        return "empty"

    # Try numeric
    numeric_count = 0
    for v in non_null[:100]:
        try:
            float(str(v).replace(",", ""))
            numeric_count += 1
        except (ValueError, TypeError):
            pass

    if numeric_count > len(non_null[:100]) * 0.8:
        return "numeric"

    return "string"


def parse_csv(file_path):
    """Parse CSV file into list of dicts."""
    rows = []
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        lines = f.readlines()

    if not lines:
        return [], []

    # Parse header
    header = [h.strip().strip('"') for h in lines[0].strip().split(',')]

    # Parse rows
    for line in lines[1:]:
        if line.strip():
            values = [v.strip().strip('"') for v in line.strip().split(',')]
            row = dict(zip(header, values))
            rows.append(row)

    return header, rows


def parse_json(file_path):
    """Parse JSON file into list of dicts."""
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if isinstance(data, list) and data:
        header = list(data[0].keys()) if isinstance(data[0], dict) else []
        return header, data
    elif isinstance(data, dict):
        return list(data.keys()), [data]

    return [], []


def parse_excel(file_path):
    """Parse Excel file into list of dicts. Requires openpyxl."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return None, "Error: openpyxl not installed. Run: pip install openpyxl"

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)

    if not header_row:
        return [], []

    header = [str(h) if h else f"col_{i}" for i, h in enumerate(header_row)]
    rows = []

    for row in rows_iter:
        row_dict = dict(zip(header, row))
        rows.append(row_dict)

    wb.close()
    return header, rows


def calculate_stats(values):
    """Calculate basic statistics for numeric values."""
    nums = []
    for v in values:
        try:
            nums.append(float(str(v).replace(",", "")))
        except (ValueError, TypeError):
            pass

    if not nums:
        return None

    nums.sort()
    n = len(nums)
    mean = sum(nums) / n

    # Calculate standard deviation
    variance = sum((x - mean) ** 2 for x in nums) / n
    std = variance ** 0.5

    return {
        "count": n,
        "min": round(nums[0], 2),
        "max": round(nums[-1], 2),
        "mean": round(mean, 2),
        "std": round(std, 2),
        "median": round(nums[n // 2], 2) if n % 2 == 1 else round((nums[n // 2 - 1] + nums[n // 2]) / 2, 2)
    }


def analyze(file_path):
    """Analyze data file and return summary."""
    path = Path(file_path)

    if not path.exists():
        return f"Error: File not found: {file_path}"

    # Parse file
    ext = path.suffix.lower()
    if ext == '.csv':
        header, rows = parse_csv(file_path)
    elif ext == '.json':
        header, rows = parse_json(file_path)
    elif ext in ('.xlsx', '.xlsm'):
        result = parse_excel(file_path)
        if isinstance(result[1], str):  # Error message
            return result[1]
        header, rows = result
    else:
        return f"Error: Unsupported file type: {ext} (use .csv, .json, or .xlsx)"

    if not rows:
        return "Error: No data found in file"

    # Build summary
    output = []
    output.append(f"## Data Summary: {path.name}")
    output.append(f"- **Rows:** {len(rows)}")
    output.append(f"- **Columns:** {len(header)}")
    output.append("")

    # Column analysis
    output.append("## Column Analysis")
    output.append("")
    output.append("| Column | Type | Missing | Stats |")
    output.append("|--------|------|---------|-------|")

    for col in header:
        values = [row.get(col) for row in rows]
        col_type = detect_type(values)
        missing = sum(1 for v in values if v is None or v == "")
        missing_pct = f"{missing} ({100*missing/len(values):.1f}%)" if values else "N/A"

        stats_str = "-"
        if col_type == "numeric":
            stats = calculate_stats(values)
            if stats:
                stats_str = f"min={stats['min']}, max={stats['max']}, mean={stats['mean']}, std={stats['std']}"

        output.append(f"| {col} | {col_type} | {missing_pct} | {stats_str} |")

    # Sample rows
    output.append("")
    output.append("## Sample Data (first 5 rows)")
    output.append("")

    sample = rows[:5]
    if sample:
        output.append("| " + " | ".join(header) + " |")
        output.append("|" + "|".join(["---"] * len(header)) + "|")
        for row in sample:
            vals = [str(row.get(col, ""))[:30] for col in header]
            output.append("| " + " | ".join(vals) + " |")

    return "\n".join(output)


def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_data.py <file-path>", file=sys.stderr)
        sys.exit(1)

    result = analyze(sys.argv[1])
    print(result)


if __name__ == "__main__":
    main()
